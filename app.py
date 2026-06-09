from flask import Flask, render_template, request, jsonify, redirect, session, send_file
from werkzeug.exceptions import RequestEntityTooLarge
from urllib.parse import urlencode
from database import (connect_db, add_student, get_students, mark_attendance, 
                      get_students_by_program_year, get_students_by_section, 
                      get_attendance_by_section, get_student_attendance_summary,
                      process_classroom_photo, get_student_face_image, get_student_statistics,
                      delete_student, get_attendance_by_program_year, delete_attendance_sheet,
                      get_attendance_sheet, get_student_by_id, update_student, get_student_voice_data,
                      get_daily_attendance, get_monthly_attendance_percentage,
                      get_attendance_summary_range, get_all_attendance_for_export,
                      get_student_monthly_attendance)
import os
from datetime import datetime
import face_recognition
import numpy as np
import cv2
from io import BytesIO
from PIL import Image
import base64
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email_utils import send_email, is_email_configured
from database import get_student_notification_data, get_students_notification_data

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
# Limit uploaded file size to 16 MB to avoid "Request Entity Too Large" errors
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024


@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(error):
    error_msg = 'Uploaded file is too large. Maximum allowed size is 16 MB.'
    if request.is_json:
        return jsonify({'error': error_msg}), 413
    return render_template('error.html', error=error_msg), 413


def normalize_time_for_db(time_value):
    if not time_value:
        return time_value
    return str(time_value)[:5]


def send_registration_email(student):
    if not student or not student.get('email'):
        return False, 'Student email is missing'

    website_url = (student.get('website_url') or request.url_root).rstrip('/')
    subject = 'Welcome to FRAS - Registration Successful'
    
    html_body = f"""
    <html>
        <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <!-- Header -->
                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px 20px; border-radius: 8px 8px 0 0; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 28px;">Welcome to FRAS!</h1>
                    <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0 0;">Face Recognition Attendance System</p>
                </div>
                
                <!-- Content -->
                <div style="background-color: #f9f9f9; padding: 30px; border: 1px solid #ddd; border-radius: 0 0 8px 8px;">
                    <p style="font-size: 16px;">Hello <strong>{student['name']}</strong>,</p>
                    
                    <p style="margin-top: 20px; font-size: 15px;">
                        Congratulations! Your face has been successfully registered in our system. 
                        You can now access the attendance system using your face recognition credentials.
                    </p>

                    <!-- Website Link -->
                    <div style="background-color: #eef2ff; border-left: 4px solid #667eea; padding: 20px; margin: 25px 0; border-radius: 4px;">
                        <h3 style="margin-top: 0; color: #667eea;">Website Link:</h3>
                        <p style="margin: 0 0 16px 0; font-size: 15px;">
                            Use the link below to open the FRAS website:
                        </p>
                        <a href="{website_url}" style="display: inline-block; background: #667eea; color: white; padding: 12px 18px; border-radius: 4px; text-decoration: none; font-weight: bold;">
                            Open FRAS Website
                        </a>
                        <p style="margin: 14px 0 0 0; font-size: 13px; color: #555;">
                            {website_url}
                        </p>
                    </div>
                    
                    <!-- Registration Details Box -->
                    <div style="background-color: white; border-left: 4px solid #667eea; padding: 20px; margin: 25px 0; border-radius: 4px;">
                        <h3 style="margin-top: 0; color: #667eea;">Your Registration Details:</h3>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="padding: 8px 0; font-weight: bold; color: #555;">Student ID:</td>
                                <td style="padding: 8px 0; color: #333;">{student['student_id']}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; font-weight: bold; color: #555;">Name:</td>
                                <td style="padding: 8px 0; color: #333;">{student['name']}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; font-weight: bold; color: #555;">Program:</td>
                                <td style="padding: 8px 0; color: #333;">{student['program']}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; font-weight: bold; color: #555;">Year:</td>
                                <td style="padding: 8px 0; color: #333;">{student['year']}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; font-weight: bold; color: #555;">Section:</td>
                                <td style="padding: 8px 0; color: #333;">{student['section']}</td>
                            </tr>
                            {f'<tr><td style="padding: 8px 0; font-weight: bold; color: #555;">Phone:</td><td style="padding: 8px 0; color: #333;">{student["phone"]}</td></tr>' if student.get('phone') else ''}
                        </table>
                    </div>
                    
                    <!-- Next Steps -->
                    <div style="background-color: #e8f4f8; border-left: 4px solid #17a2b8; padding: 20px; margin: 25px 0; border-radius: 4px;">
                        <h3 style="margin-top: 0; color: #17a2b8;">What's Next?</h3>
                        <ul style="margin: 0; padding-left: 20px;">
                            <li style="margin: 10px 0;">Your face has been successfully registered</li>
                            <li style="margin: 10px 0;">You can now access the attendance system</li>
                            <li style="margin: 10px 0;">Your attendance will be automatically recorded via facial recognition</li>
                            <li style="margin: 10px 0;">Check the dashboard regularly to view your attendance reports</li>
                        </ul>
                    </div>
                    
                    <!-- Support -->
                    <p style="margin: 25px 0 0 0; color: #666; font-size: 14px;">
                        If you have any questions or need assistance, please contact the administration office.
                    </p>
                    
                    <p style="margin: 15px 0; color: #666; font-size: 14px;">
                        <strong>Note:</strong> This is an automated message. Please do not reply to this email.
                    </p>
                </div>
                
                <!-- Footer -->
                <div style="background-color: #333; color: white; padding: 20px; text-align: center; font-size: 12px; border-radius: 0;">
                    <p style="margin: 0;">Face Recognition Attendance System</p>
                    <p style="margin: 5px 0 0 0; opacity: 0.8;">© 2024. All rights reserved.</p>
                </div>
            </div>
        </body>
    </html>
    """
    
    sent, error = send_email(student['email'], subject, html_body, is_html=True)
    if not sent:
        print(f"Registration email not sent to {student['email']}: {error}")
        return False, error
    return True, None


def send_attendance_emails(student_rows, attendance_date, attendance_time, subject_name, period_duration, status_by_student_id):
    if not is_email_configured():
        return

    subject_label = subject_name or 'Attendance Update'
    period_label = period_duration or '-'

    for row in student_rows:
        student_id, name, program, year, section, email = row
        if not email:
            continue

        status = status_by_student_id.get(student_id)
        if not status:
            continue

        message = (
            f"Hello {name},\n\n"
            f"Your attendance has been recorded.\n\n"
            f"Student ID: {student_id}\n"
            f"Program: {program}\n"
            f"Year: {year}\n"
            f"Section: {section}\n"
            f"Date: {attendance_date}\n"
            f"Time: {attendance_time}\n"
            f"Subject: {subject_label}\n"
            f"Period: {period_label}\n"
            f"Status: {status}\n"
        )
        sent, error = send_email(email, f'Attendance for {subject_label}', message, is_html=False)
        if not sent:
            print(f"Attendance email not sent to {email}: {error}")


def format_period_time(time_24h):
    """Convert HH:MM to h:mm am/pm."""
    try:
        return datetime.strptime(time_24h, '%H:%M').strftime('%I:%M %p').lstrip('0').lower()
    except Exception:
        return time_24h


def normalize_speech_text(value):
    if not value:
        return ''
    chars = []
    for ch in value.lower().strip():
        if ch.isalnum() or ch.isspace():
            chars.append(ch)
        else:
            chars.append(' ')
    return ' '.join(''.join(chars).split())


# Define available programs, years, and sections
PROGRAMS = {
    'MCA': {
        'years': [1, 2],
        'sections': ['A', 'B']
    },
    'BCA': {
        'years': [1, 2, 3],
        'sections': ['A', 'B']
    }
}

# Home route
@app.route('/')
def index():
    return redirect('/admin_login')


# Admin login route
@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        # Simple authentication (should be enhanced with proper password hashing)
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Demo credentials
        if username == 'admin' and password == 'admin123':
            session['admin'] = True
            return redirect('/dashboard')
        else:
            return render_template('admin_login.html', error='Invalid username or password')
    return render_template('admin_login.html')


# Student registration route (public access)
@app.route('/student_register', methods=['GET', 'POST'])
def student_register():
    # Get URL parameters for pre-filling the form
    selected_program = request.args.get('program')
    selected_year = request.args.get('year')
    selected_section = request.args.get('section')
    
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        name = request.form.get('name')
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        voice_passphrase = request.form.get('voice_passphrase', '').strip()
        face_image_data = request.form.get('face_image')
        
        # Preserve submitted class selection if validation fails.
        selected_program = program or selected_program
        selected_year = year or selected_year
        selected_section = section or selected_section
        
        # if webcam capture not used, check for uploaded file
        if not face_image_data and 'photoFile' in request.files:
            file = request.files['photoFile']
            if file and file.filename:
                try:
                    img_bytes = file.read()
                    encoded = base64.b64encode(img_bytes).decode('utf-8')
                    face_image_data = 'data:image/jpeg;base64,' + encoded
                except Exception as e:
                    print(f"Error reading uploaded photo: {e}")
        
        # Validate inputs
        if not all([student_id, name, program, year, section]):
            return render_template('student_register.html', 
                                 error='All required fields must be filled', 
                                 programs=PROGRAMS,
                                 selected_program=selected_program,
                                 selected_year=selected_year,
                                 selected_section=selected_section)
        
        if not face_image_data:
            return render_template('student_register.html', 
                                 error='Please capture or upload your face photo before registering', 
                                 programs=PROGRAMS,
                                 selected_program=selected_program,
                                 selected_year=selected_year,
                                 selected_section=selected_section)
        
        try:
            year = int(year)
            
            # Process face image
            face_image_binary = None
            face_encoding_binary = None
            
            try:
                # Decode base64 image
                face_data = face_image_data.split(',')[1]
                face_image_binary = base64.b64decode(face_data)
                
                # Convert to numpy array for face recognition
                nparr = np.frombuffer(face_image_binary, np.uint8)
                img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                
                # Get face encoding
                rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                face_locations = face_recognition.face_locations(rgb_img)
                
                if face_locations:
                    face_encodings = face_recognition.face_encodings(rgb_img, face_locations)
                    if face_encodings:
                        face_encoding_binary = face_encodings[0].tobytes()
                        
            except Exception as e:
                print(f"Face processing error: {e}")
                return render_template('student_register.html', error=f'Error processing photo: {str(e)}', programs=PROGRAMS, selected_program=selected_program, selected_year=selected_year, selected_section=selected_section)
            
            if not face_encoding_binary:
                return render_template('student_register.html', error='Face encoding failed. Please try a clearer photo.', programs=PROGRAMS, selected_program=selected_program, selected_year=selected_year, selected_section=selected_section)

            success = add_student(student_id, name, program, year, section, email, phone, face_image_binary, face_encoding_binary, voice_passphrase)
            
            if success:
                email_sent, email_error = send_registration_email({
                    'student_id': student_id,
                    'name': name,
                    'program': program.upper(),
                    'year': year,
                    'section': section.upper(),
                    'email': email,
                    'phone': phone,
                    'website_url': request.url_root
                })
                email_warning = None if email_sent else f'Registration saved, but welcome email was not sent: {email_error}'
                return render_template('student_register.html', 
                                     success='Registration successful!', 
                                     email_warning=email_warning,
                                     programs=PROGRAMS,
                                     selected_program=selected_program,
                                     selected_year=selected_year,
                                     selected_section=selected_section)
            else:
                return render_template('student_register.html', 
                                     error='Registration failed. Student ID may already exist.', 
                                     programs=PROGRAMS,
                                     selected_program=selected_program,
                                     selected_year=selected_year,
                                     selected_section=selected_section)
                                     
        except Exception as e:
            return render_template('student_register.html', 
                                 error=f'Error: {str(e)}', 
                                 programs=PROGRAMS,
                                 selected_program=selected_program,
                                 selected_year=selected_year,
                                 selected_section=selected_section)
    
    return render_template('student_register.html', 
                         programs=PROGRAMS,
                         selected_program=selected_program,
                         selected_year=selected_year,
                         selected_section=selected_section)


# Generate registration links for students
@app.route('/generate_links', methods=['GET', 'POST'])
def generate_links():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    links = []
    
    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        
        if program and year and section:
            try:
                year = int(year)
                base_url = request.host_url.rstrip('/')
                registration_link = f"{base_url}/student_register"
                
                # Generate a class-specific link (could include parameters)
                class_link = f"{registration_link}?program={program}&year={year}&section={section}"
                
                links.append({
                    'program': program,
                    'year': year,
                    'section': section,
                    'link': class_link,
                    'description': f'{program} Year {year} Section {section}'
                })
                
            except Exception as e:
                print(f"Error generating links: {e}")
    
    return render_template('generate_links.html', links=links, programs=PROGRAMS)


# Dashboard route
@app.route('/dashboard')
def dashboard():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    students = get_students()
    stats = get_student_statistics()
    
    # Organize statistics by program
    organized_stats = {}
    total_students = 0
    
    for program, year, section, count in stats:
        if program not in organized_stats:
            organized_stats[program] = {}
        if year not in organized_stats[program]:
            organized_stats[program][year] = {}
        organized_stats[program][year][section] = count
        total_students += count
    
    # build quick photo-attendance links with program and year preselected
    photo_links = []
    base_url = request.host_url.rstrip('/')
    for prog, info in PROGRAMS.items():
        for y in info['years']:
            photo_links.append({
                'program': prog,
                'year': y,
                'link': f"{base_url}/photo_attendance?program={prog}&year={y}"
            })
    
    return render_template('dashboard.html', 
                         students=students, 
                         stats=organized_stats, 
                         total_students=total_students, 
                         programs=PROGRAMS,
                         photo_links=photo_links)


# Student statistics route
@app.route('/student_statistics')
def student_statistics():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    stats = get_student_statistics()
    
    # Organize statistics by program
    organized_stats = {}
    total_students = 0
    
    for program, year, section, count in stats:
        if program not in organized_stats:
            organized_stats[program] = {}
        if year not in organized_stats[program]:
            organized_stats[program][year] = {}
        organized_stats[program][year][section] = count
        total_students += count
    
    return render_template('student_statistics.html', 
                         stats=organized_stats, 
                         total_students=total_students, 
                         programs=PROGRAMS)


# Register student route
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        name = request.form.get('name')
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        voice_passphrase = request.form.get('voice_passphrase', '').strip()
        face_image_data = request.form.get('face_image')
        
        # if webcam capture not used, check for uploaded file
        if not face_image_data and 'photoFile' in request.files:
            file = request.files['photoFile']
            if file and file.filename:
                try:
                    img_bytes = file.read()
                    encoded = base64.b64encode(img_bytes).decode('utf-8')
                    face_image_data = 'data:image/jpeg;base64,' + encoded
                except Exception as e:
                    print(f"Error reading uploaded photo: {e}")
        
        # Validate inputs
        if not all([student_id, name, program, year, section]):
            return render_template('register.html', 
                                 error='All required fields must be filled', 
                                 programs=PROGRAMS)
        
        if not face_image_data:
            return render_template('register.html', 
                                 error='Please capture or upload your face photo before registering', 
                                 programs=PROGRAMS)
        
        try:
            year = int(year)
            
            # Process face image
            face_image_binary = None
            face_encoding_binary = None
            
            try:
                # Decode base64 image
                face_data = face_image_data.split(',')[1]
                face_image_binary = base64.b64decode(face_data)
                
                # Convert to numpy array for face recognition
                nparr = np.frombuffer(face_image_binary, np.uint8)
                img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                
                # Get face encoding
                rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                face_locations = face_recognition.face_locations(rgb_img)
                
                if face_locations:
                    # Improve accuracy: resample face 10 times to create a robust encoding
                    face_encodings = face_recognition.face_encodings(rgb_img, face_locations, num_jitters=10)
                    if face_encodings:
                        face_encoding_binary = face_encodings[0].tobytes()
                    else:
                        # Image valid, but no face found by library
                        return render_template('register.html', error='No face detected in the photo. Please try again with better lighting.', programs=PROGRAMS)
                        
            except Exception as e:
                print(f"Face processing error: {e}")
                return render_template('register.html', error=f'Error processing photo: {str(e)}', programs=PROGRAMS)
            
            if not face_encoding_binary:
                return render_template('register.html', error='Face encoding failed. Please try a clearer photo.', programs=PROGRAMS)

            success = add_student(student_id, name, program, year, section, email, phone, face_image_binary, face_encoding_binary, voice_passphrase)
            if success:
                email_sent, email_error = send_registration_email({
                    'student_id': student_id,
                    'name': name,
                    'program': program.upper(),
                    'year': year,
                    'section': section.upper(),
                    'email': email,
                    'phone': phone,
                    'website_url': request.url_root
                })
                email_warning = None if email_sent else f'Student saved, but welcome email was not sent: {email_error}'
                return render_template('register.html', 
                                     success='✓ Student registered successfully with face data!', 
                                     email_warning=email_warning,
                                     programs=PROGRAMS)
            return render_template('register.html',
                                 error='Registration failed. Student ID may already exist.',
                                 programs=PROGRAMS)
        except Exception as e:
            return render_template('register.html', 
                                 error=f'Error: {str(e)}', 
                                 programs=PROGRAMS)
    
    return render_template('register.html', programs=PROGRAMS)


# Photo attendance marking page
@app.route('/photo_attendance', methods=['GET', 'POST'])
def photo_attendance():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    # Get URL parameters for pre-selecting form values
    selected_program = request.args.get('program')
    selected_year = request.args.get('year')
    selected_section = request.args.get('section')
    
    # Convert year to int if provided
    if selected_year:
        try:
            selected_year = int(selected_year)
        except ValueError:
            selected_year = None
    
    # Get today's date/time for default values
    now = datetime.now()
    today_date = now.strftime('%Y-%m-%d')
    today_time = now.strftime('%H:%M')
    
    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        date = request.form.get('date')
        selected_time = request.form.get('time') or today_time
        subject_name = request.form.get('subject_name', '').strip()
        period_from = request.form.get('period_from', '').strip()
        period_to = request.form.get('period_to', '').strip()
        classroom_photo = request.form.get('classroom_photo')
        selected_program = program
        selected_section = section
        try:
            selected_year = int(year) if year else None
        except ValueError:
            selected_year = year
        period_duration = ''
        if period_from and period_to:
            period_duration = f"from {format_period_time(period_from)} to {format_period_time(period_to)}"
        if period_from and period_to and period_from >= period_to:
            return render_template('photo_attendance.html',
                                 error='Period "to" time must be later than "from" time',
                                 programs=PROGRAMS,
                                 selected_program=selected_program,
                                 selected_year=selected_year,
                                 selected_section=selected_section,
                                 selected_subject_name=subject_name,
                                 selected_period_duration=period_duration,
                                 selected_period_from=period_from,
                                 selected_period_to=period_to,
                                 selected_date=date,
                                 selected_time=selected_time,
                                 today_time=today_time,
                                 today_date=today_date)
        
        # Handle standard file upload if the hidden base64 field is missing.
        if not classroom_photo and (
            'classroom_photo_file' in request.files or 'classroom_photo' in request.files
        ):
            file = request.files.get('classroom_photo_file') or request.files.get('classroom_photo')
            if file and file.filename:
                try:
                    img_bytes = file.read()
                    encoded = base64.b64encode(img_bytes).decode('utf-8')
                    mimetype = file.mimetype or 'image/jpeg'
                    classroom_photo = f'data:{mimetype};base64,{encoded}'
                except Exception as e:
                    print(f"Error reading uploaded photo: {e}")

        if not all([program, year, section, date, selected_time, subject_name, period_from, period_to, classroom_photo]):
            return render_template('photo_attendance.html', 
                                 error='All fields are required', 
                                 programs=PROGRAMS,
                                 selected_program=selected_program,
                                 selected_year=selected_year,
                                 selected_section=selected_section,
                                 selected_subject_name=subject_name,
                                 selected_period_duration=period_duration,
                                 selected_period_from=period_from,
                                 selected_period_to=period_to,
                                 selected_date=date,
                                 selected_time=selected_time,
                                 today_time=today_time,
                                 today_date=today_date)
        
        try:
            year = int(year)
            # Parse the date
            attendance_date = datetime.strptime(date, '%Y-%m-%d').date()
            
            result = process_classroom_photo(
                classroom_photo, program, year, section, attendance_date, selected_time,
                subject_name, period_duration
            )
            
            if 'error' in result:
                return render_template('photo_attendance.html', 
                                     error=result['error'], 
                                     programs=PROGRAMS,
                                     selected_program=selected_program,
                                     selected_year=selected_year,
                                     selected_section=selected_section,
                                     selected_subject_name=subject_name,
                                     selected_period_duration=period_duration,
                                     selected_period_from=period_from,
                                     selected_period_to=period_to,
                                     selected_date=date,
                                     selected_time=selected_time,
                                     today_time=today_time,
                                     today_date=today_date)

            status_by_student_id = {}
            for student_id in result.get('present_students', []):
                status_by_student_id[student_id] = 'Present'
            for student_id in result.get('absent_students', []):
                status_by_student_id[student_id] = 'Absent'

            send_attendance_emails(
                get_students_notification_data(list(status_by_student_id.keys())),
                date,
                selected_time,
                subject_name,
                period_duration,
                status_by_student_id
            )

            student_names = {student[0]: student[1] for student in get_students_by_section(program, year, section)}
            records = [
                {
                    'student_id': item['student_id'],
                    'name': student_names.get(item['student_id'], item['student_id']),
                    'status': item['status']
                }
                for item in result.get('attendance', [])
            ]
            return render_template(
                'attendance_result.html',
                source='photo',
                program=program,
                year=year,
                section=section,
                date=date,
                selected_time=selected_time,
                subject_name=subject_name,
                period_duration=period_duration,
                faces_detected=result.get('total_faces_detected', 0),
                present_count=result.get('present_count', 0),
                absent_count=result.get('absent_count', 0),
                records=records
            )
                                 
        except Exception as e:
            return render_template('photo_attendance.html', 
                                 error=f'Error processing photo: {str(e)}', 
                                 programs=PROGRAMS,
                                 selected_program=selected_program,
                                 selected_year=selected_year,
                                 selected_section=selected_section,
                                 selected_subject_name=subject_name,
                                 selected_period_duration=period_duration,
                                 selected_period_from=period_from,
                                 selected_period_to=period_to,
                                 selected_date=date,
                                 selected_time=selected_time,
                                 today_time=today_time,
                                 today_date=today_date)
    
    return render_template('photo_attendance.html', 
                         programs=PROGRAMS,
                         selected_program=selected_program,
                         selected_year=selected_year,
                         selected_section=selected_section,
                         today_time=today_time,
                         today_date=today_date)


# Attendance marking page
@app.route('/mark_attendance', methods=['GET', 'POST'])
def mark_attendance_page():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    # Get today's date/time for default values
    now = datetime.now()
    today_date = now.strftime('%Y-%m-%d')
    today_time = now.strftime('%H:%M')
    
    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        date = request.form.get('date')
        selected_time = request.form.get('time') or today_time
        subject_name = request.form.get('subject_name', '').strip()
        period_from = request.form.get('period_from', '').strip()
        period_to = request.form.get('period_to', '').strip()
        period_duration = ''
        if period_from and period_to:
            period_duration = f"from {format_period_time(period_from)} to {format_period_time(period_to)}"
        if period_from and period_to and period_from >= period_to:
            return render_template('mark_attendance.html',
                                 error='Period "to" time must be later than "from" time',
                                 programs=PROGRAMS,
                                 selected_program=program,
                                 selected_year=year,
                                 selected_section=section,
                                 selected_date=date,
                                 selected_subject_name=subject_name,
                                 selected_period_duration=period_duration,
                                 selected_period_from=period_from,
                                 selected_period_to=period_to,
                                 selected_time=selected_time,
                                 today_time=today_time,
                                 today_date=today_date)
        
        try:
            year = int(year)
            students = get_students_by_section(program, year, section)
            return render_template('mark_attendance.html', 
                                 students=students, 
                                 programs=PROGRAMS,
                                 selected_program=program,
                                 selected_year=year,
                                 selected_section=section,
                                 selected_date=date,
                                 selected_subject_name=subject_name,
                                 selected_period_duration=period_duration,
                                 selected_period_from=period_from,
                                 selected_period_to=period_to,
                                 selected_time=selected_time,
                                 today_time=today_time,
                                 today_date=today_date)
        except Exception as e:
            return render_template('mark_attendance.html', 
                                 error=f'Error: {str(e)}', 
                                 programs=PROGRAMS,
                                 today_time=today_time,
                                 today_date=today_date)
    
    return render_template('mark_attendance.html', 
                         programs=PROGRAMS,
                         today_time=today_time,
                         today_date=today_date)


# API endpoint to mark attendance
@app.route('/api/mark_attendance', methods=['POST'])
def api_mark_attendance():
    if 'admin' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        student_id = request.json.get('student_id')
        status = request.json.get('status', 'Present')
        date = request.json.get('date')  # Optional date parameter
        time = request.json.get('time')  # Optional time parameter
        subject_name = request.json.get('subject_name')
        period_duration = request.json.get('period_duration')
        
        if mark_attendance(student_id, status, date, time, subject_name, period_duration):
            student_row = get_student_notification_data(student_id)
            send_attendance_emails(
                [student_row] if student_row else [],
                date,
                time,
                subject_name,
                period_duration,
                {student_id: status}
            )
            return jsonify({'success': True, 'message': f'Attendance marked as {status}'})
        else:
            return jsonify({'success': False, 'message': 'Failed to mark attendance'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/voice_checkin', methods=['GET', 'POST'])
def voice_checkin():
    if 'admin' not in session:
        return redirect('/admin_login')

    now = datetime.now()
    today_date = now.strftime('%Y-%m-%d')
    today_time = now.strftime('%H:%M')

    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        date = request.form.get('date')
        selected_time = request.form.get('time') or today_time
        subject_name = request.form.get('subject_name', '').strip()
        period_from = request.form.get('period_from', '').strip()
        period_to = request.form.get('period_to', '').strip()
        period_duration = ''
        if period_from and period_to:
            period_duration = f"from {format_period_time(period_from)} to {format_period_time(period_to)}"

        return render_template(
            'voice_checkin.html',
            programs=PROGRAMS,
            selected_program=program,
            selected_year=year,
            selected_section=section,
            selected_date=date,
            selected_time=selected_time,
            selected_subject_name=subject_name,
            selected_period_from=period_from,
            selected_period_to=period_to,
            selected_period_duration=period_duration,
            today_date=today_date,
            today_time=today_time
        )

    return render_template('voice_checkin.html', programs=PROGRAMS, today_date=today_date, today_time=today_time)


@app.route('/api/voice_checkin', methods=['POST'])
def api_voice_checkin():
    if 'admin' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    try:
        student_id = (request.json.get('student_id') or '').strip()
        spoken_text = (request.json.get('spoken_text') or '').strip()
        program = (request.json.get('program') or '').strip().upper()
        year = request.json.get('year')
        section = (request.json.get('section') or '').strip().upper()
        date = request.json.get('date')
        selected_time = request.json.get('time')
        subject_name = (request.json.get('subject_name') or '').strip()
        period_duration = (request.json.get('period_duration') or '').strip()

        if not all([student_id, spoken_text, program, year, section, date, selected_time, subject_name, period_duration]):
            return jsonify({'success': False, 'message': 'All fields are required'}), 400

        student = get_student_voice_data(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # student tuple: id, name, voice_passphrase, program, year, section
        if str(student[3]).strip().upper() != program or int(student[4]) != int(year) or str(student[5]).strip().upper() != section:
            return jsonify({'success': False, 'message': 'Student is not in selected class'}), 400

        expected_phrase = normalize_speech_text(student[2] or '')
        actual_spoken = normalize_speech_text(spoken_text)
        if not expected_phrase:
            return jsonify({'success': False, 'message': 'No voice passphrase set for this student'}), 400
        if expected_phrase not in actual_spoken:
            return jsonify({'success': False, 'message': 'Voice passphrase did not match'}), 400

        if not mark_attendance(student_id, "Present", date, selected_time, subject_name, period_duration):
            return jsonify({'success': False, 'message': 'Failed to mark attendance'}), 500

        student_row = get_student_notification_data(student_id)
        send_attendance_emails(
            [student_row] if student_row else [],
            date,
            selected_time,
            subject_name,
            period_duration,
            {student_id: 'Present'}
        )

        query = urlencode({
            'source': 'manual',
            'program': program,
            'year': year,
            'section': section,
            'date': date,
            'time': selected_time,
            'subject_name': subject_name,
            'period_duration': period_duration
        })
        return jsonify({'success': True, 'message': f'Attendance marked for {student[1]}', 'redirect_url': f"/attendance_result?{query}"})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/attendance_result')
def attendance_result():
    if 'admin' not in session:
        return redirect('/admin_login')

    source = request.args.get('source', 'manual')
    program = (request.args.get('program') or '').strip().upper()
    year = request.args.get('year')
    section = (request.args.get('section') or '').strip().upper()
    date = request.args.get('date')
    selected_time = request.args.get('time')
    subject_name = request.args.get('subject_name') or '-'
    period_duration = request.args.get('period_duration') or '-'
    faces_detected = int(request.args.get('faces_detected', 0) or 0)

    records = []
    present_count = 0
    absent_count = 0

    try:
        year_int = int(year)
        raw_attendance = get_attendance_sheet(
            program,
            year_int,
            section,
            date,
            normalize_time_for_db(selected_time),
            None if subject_name == '-' else subject_name,
            None if period_duration == '-' else period_duration
        )

        for row in raw_attendance:
            status = row[4]
            if status == 'Present':
                present_count += 1
            elif status == 'Absent':
                absent_count += 1

            records.append({
                'student_id': row[1],
                'name': row[7],
                'status': status
            })

    except Exception as e:
        print(f"Error loading attendance result: {e}")

    return render_template(
        'attendance_result.html',
        source=source,
        program=program,
        year=year,
        section=section,
        date=date,
        selected_time=selected_time,
        subject_name=subject_name,
        period_duration=period_duration,
        faces_detected=faces_detected,
        present_count=present_count,
        absent_count=absent_count,
        records=records
    )


# Attendance mark sheet page
@app.route('/attendance_marksheet', methods=['GET', 'POST'])
@app.route('/report', methods=['GET', 'POST'])
def report():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    attendance_sheets = []
    selected_program = None
    selected_year = None
    selected_section = None

    # Support both POST (form submit) and GET (redirect after delete)
    program = request.values.get('program')
    year = request.values.get('year')
    section = request.values.get('section')

    if program and year:
        try:
            program = program.strip().upper()
            if program not in PROGRAMS:
                raise ValueError("Invalid program")

            year = int(year)
            selected_program = program
            selected_year = year
            if section and section in PROGRAMS.get(program, {}).get('sections', []):
                selected_section = section
        except Exception as e:
            print(f"Error parsing report filters: {e}")
            year = None

        if year is not None:
            try:
                # Get attendance sheet details (with date and time), grouped by section/date/time
                raw_attendance = get_attendance_by_program_year(program, year)
                sheet_map = {}

                for row in raw_attendance:
                    # row: attendance_id, student_id, date, time, status, subject_name, period_duration, name, program, year, section
                    row_program = str(row[8]).strip().upper()
                    if row_program != selected_program:
                        continue

                    attendance_date = row[2]
                    attendance_time = row[3]
                    row_subject_name = row[5] or '-'
                    row_period_duration = row[6] or '-'
                    row_section = row[10]

                    # Optional section filter to view only one section's mark sheets.
                    if selected_section and row_section != selected_section:
                        continue

                    date_str = attendance_date.strftime('%Y-%m-%d') if hasattr(attendance_date, 'strftime') else str(attendance_date)
                    time_str = attendance_time.strftime('%H:%M:%S') if hasattr(attendance_time, 'strftime') else str(attendance_time)
                    group_key = (row_section, date_str, time_str, row_subject_name, row_period_duration)

                    if group_key not in sheet_map:
                        sheet_map[group_key] = {
                            'section': row_section,
                            'date': date_str,
                            'time': time_str,
                            'subject_name': row_subject_name,
                            'period_duration': row_period_duration,
                            'records': [],
                            'present_count': 0,
                            'absent_count': 0
                        }

                    status = row[4]
                    if status == 'Present':
                        sheet_map[group_key]['present_count'] += 1
                    elif status == 'Absent':
                        sheet_map[group_key]['absent_count'] += 1

                    sheet_map[group_key]['records'].append({
                        'student_id': row[1],
                        'name': row[7],
                        'status': status
                    })

                attendance_sheets = list(sheet_map.values())
            except Exception as e:
                print(f"Error fetching attendance sheet data: {e}")
    
    return render_template('report.html', 
                         attendance_sheets=attendance_sheets,
                         programs=PROGRAMS,
                         selected_program=selected_program,
                         selected_year=selected_year,
                         selected_section=selected_section)


@app.route('/delete_attendance_sheet', methods=['POST'])
def delete_attendance_sheet_route():
    if 'admin' not in session:
        return redirect('/admin_login')

    program = request.form.get('program')
    year = request.form.get('year')
    section = request.form.get('section')
    date = request.form.get('date')
    time = request.form.get('time')
    subject_name = request.form.get('subject_name')
    period_duration = request.form.get('period_duration')
    selected_section_filter = request.form.get('selected_section_filter')

    if all([program, year, section, date, time]):
        try:
            delete_attendance_sheet(program, int(year), section, date, time, subject_name, period_duration)
        except Exception as e:
            print(f"Error deleting attendance sheet: {e}")

    # Return to attendance mark sheet page with same selected filters
    redirect_url = f"/attendance_marksheet?program={program}&year={year}"
    if selected_section_filter:
        redirect_url += f"&section={selected_section_filter}"
    return redirect(redirect_url)


# Student statistics
@app.route('/student_stats/<student_id>')
def student_stats(student_id):
    if 'admin' not in session:
        return redirect('/admin_login')
    
    try:
        attendance = get_student_attendance_summary(student_id)
        present_count = sum(1 for record in attendance if record[4] == 'Present')
        absent_count = sum(1 for record in attendance if record[4] == 'Absent')
        
        stats = {
            'total_days': len(attendance),
            'present_days': present_count,
            'absent_days': absent_count,
            'percentage': round((present_count / len(attendance) * 100), 2) if attendance else 0
        }
        
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# View student photos
@app.route('/view_students', methods=['GET', 'POST'])
def view_students():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    students = []
    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        
        try:
            year = int(year)
            students = get_students_by_section(program, year, section)
            return render_template('view_students.html', 
                                 students=students,
                                 programs=PROGRAMS,
                                 selected_program=program,
                                 selected_year=year,
                                 selected_section=section)
        except Exception as e:
            return render_template('view_students.html', 
                                 error=f'Error: {str(e)}',
                                 programs=PROGRAMS)
    
    return render_template('view_students.html', programs=PROGRAMS)


# Get student photo
@app.route('/api/student_photo/<student_id>')
def get_student_photo(student_id):
    if 'admin' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        photo_data = get_student_face_image(student_id)
        if photo_data:
            return jsonify({'photo': photo_data})
        return jsonify({'error': 'No photo found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# View student details by class
@app.route('/student_details', methods=['GET', 'POST'])
def student_details():
    if 'admin' not in session:
        return redirect('/admin_login')
    
    students = []
    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        
        try:
            year = int(year)
            students = get_students_by_section(program, year, section)
            return render_template('student_details.html', 
                                 students=students,
                                 programs=PROGRAMS,
                                 selected_program=program,
                                 selected_year=year,
                                 selected_section=section)
        except Exception as e:
            return render_template('student_details.html', 
                                 error=f'Error: {str(e)}',
                                 programs=PROGRAMS)
    
    return render_template('student_details.html', programs=PROGRAMS)


# Delete student route
@app.route('/delete_student/<student_id>', methods=['POST'])
def delete_student_route(student_id):
    if 'admin' not in session:
        return redirect('/admin_login')
    
    if delete_student(student_id):
        return redirect('/student_details')
    else:
        return "Error deleting student", 500


@app.route('/edit_student/<student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    if 'admin' not in session:
        return redirect('/admin_login')

    student = get_student_by_id(student_id)
    if not student:
        return "Student not found", 404

    if request.method == 'POST':
        name = request.form.get('name')
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')

        if not all([name, program, year, section]):
            return render_template('edit_student.html',
                                 error='Name, Program, Year, and Section are required.',
                                 programs=PROGRAMS,
                                 student=student)

        try:
            year = int(year)
            if update_student(student_id, name, program, year, section, email, phone):
                updated_student = get_student_by_id(student_id)
                return render_template('edit_student.html',
                                     success='Student updated successfully.',
                                     programs=PROGRAMS,
                                     student=updated_student)
            return render_template('edit_student.html',
                                 error='Failed to update student.',
                                 programs=PROGRAMS,
                                 student=student)
        except Exception as e:
            return render_template('edit_student.html',
                                 error=f'Error: {str(e)}',
                                 programs=PROGRAMS,
                                 student=student)

    return render_template('edit_student.html',
                         programs=PROGRAMS,
                         student=student)


# Daily Attendance Report Route
@app.route('/daily_report', methods=['GET', 'POST'])
def daily_report():
    if 'admin' not in session:
        return redirect('/admin_login')

    attendance_data = []
    selected_program = None
    selected_year = None
    selected_section = None
    selected_date = None
    present_count = 0
    absent_count = 0
    total_count = 0

    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        date = request.form.get('date')

        if program and year and section and date:
            try:
                program = program.strip().upper()
                if program not in PROGRAMS:
                    raise ValueError("Invalid program")
                
                year = int(year)
                selected_program = program
                selected_year = year
                selected_section = section.upper()
                selected_date = date

                attendance_data = get_daily_attendance(program, year, section.upper(), date)
                
                # Calculate statistics
                for record in attendance_data:
                    total_count += 1
                    if record[2] == 'Present':
                        present_count += 1
                    elif record[2] == 'Absent':
                        absent_count += 1
                        
            except Exception as e:
                print(f"Error fetching daily report: {e}")

    return render_template('daily_report.html',
                         attendance_data=attendance_data,
                         programs=PROGRAMS,
                         selected_program=selected_program,
                         selected_year=selected_year,
                         selected_section=selected_section,
                         selected_date=selected_date,
                         present_count=present_count,
                         absent_count=absent_count,
                         total_count=total_count)


# Monthly Attendance Percentage Report Route
@app.route('/monthly_report', methods=['GET', 'POST'])
def monthly_report():
    if 'admin' not in session:
        return redirect('/admin_login')

    attendance_data = []
    selected_program = None
    selected_year = None
    selected_section = None
    selected_month = None
    selected_year_val = None
    total_students = 0

    if request.method == 'POST':
        program = request.form.get('program')
        year = request.form.get('year')
        section = request.form.get('section')
        month = request.form.get('month')
        year_val = request.form.get('year_val')

        if program and year and section and month and year_val:
            try:
                program = program.strip().upper()
                section = section.strip().upper()
                if program not in PROGRAMS:
                    raise ValueError("Invalid program")
                
                year = int(year)
                month = int(month)
                year_val = int(year_val)
                
                selected_program = program
                selected_year = year
                selected_section = section
                selected_month = month
                selected_year_val = year_val

                attendance_data = get_monthly_attendance_percentage(program, year, month, year_val)
                attendance_data = [record for record in attendance_data if str(record[2]).strip().upper() == section]
                total_students = len(attendance_data)
                        
            except Exception as e:
                print(f"Error fetching monthly report: {e}")

    return render_template('monthly_report.html',
                         attendance_data=attendance_data,
                         programs=PROGRAMS,
                         selected_program=selected_program,
                         selected_year=selected_year,
                         selected_section=selected_section,
                         selected_month=selected_month,
                         selected_year_val=selected_year_val,
                         total_students=total_students)


# Export Attendance Report Route
@app.route('/export_report', methods=['GET', 'POST'])
def export_report():
    if 'admin' not in session:
        return redirect('/admin_login')

    if request.method == 'POST':
        export_type = request.form.get('export_type', 'csv')
        report_type = request.form.get('report_type', 'attendance')

        try:
            if report_type == 'daily':
                program = request.form.get('daily_program', '').strip().upper()
                year = request.form.get('daily_year', '')
                section = request.form.get('daily_section', '').strip().upper()
                start_date = request.form.get('daily_start_date', '')
                end_date = request.form.get('daily_end_date', '')
                year = int(year) if year else None

                if not (program and year and section and start_date):
                    return redirect('/export_report')

                # Daily report for a specific section and date range
                attendance_data = get_attendance_summary_range(program, year, section, start_date, end_date or start_date)
                
                if export_type == 'csv':
                    return export_daily_csv(attendance_data, program, year, section, start_date)
                else:
                    return export_daily_excel(attendance_data, program, year, section, start_date)
                    
            elif report_type == 'monthly':
                # Monthly report for entire program/year
                program = request.form.get('monthly_program', '').strip().upper()
                year = request.form.get('monthly_year', '')
                section = request.form.get('monthly_section', '').strip().upper()
                month = request.form.get('monthly_month', '')
                year_val = request.form.get('monthly_year_val', '')

                if not (program and year and section and month and year_val):
                    return redirect('/export_report')

                year = int(year)
                month = int(month)
                year_val = int(year_val)
                
                attendance_data = get_monthly_attendance_percentage(program, year, month, year_val)
                attendance_data = [record for record in attendance_data if str(record[2]).strip().upper() == section]
                
                if export_type == 'csv':
                    return export_monthly_csv(attendance_data, program, year, month, year_val)
                else:
                    return export_monthly_excel(attendance_data, program, year, month, year_val)
                    
            elif report_type == 'full':
                # Full attendance export
                program = request.form.get('daily_program', '').strip().upper()
                year = request.form.get('daily_year', '')
                section = request.form.get('daily_section', '').strip().upper()
                start_date = request.form.get('daily_start_date', '')
                end_date = request.form.get('daily_end_date', '')
                year = int(year) if year else None
                attendance_data = get_all_attendance_for_export(program or None, year, section or None, start_date or None, end_date or None)
                
                if export_type == 'csv':
                    return export_full_csv(attendance_data)
                else:
                    return export_full_excel(attendance_data)
                    
        except Exception as e:
            print(f"Error exporting report: {e}")
            return redirect(f'/monthly_report?error=Export failed: {str(e)}')

    return render_template('export_report.html', programs=PROGRAMS)


# Helper functions for CSV/Excel export
def export_daily_csv(attendance_data, program, year, section, date):
    """Export daily attendance to CSV"""
    output = BytesIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Daily Attendance Report'])
    writer.writerow([f'Program: {program}', f'Year: {year}', f'Section: {section}', f'Date: {date}'])
    writer.writerow([])
    
    writer.writerow(['Student ID', 'Name', 'Status', 'Time', 'Subject', 'Period'])
    
    # Write data
    for record in attendance_data:
        student_id, name, status, att_date, att_time, subject, period = record
        status = status or 'No Record'
        time_str = att_time.strftime('%H:%M:%S') if hasattr(att_time, 'strftime') else str(att_time or '')
        writer.writerow([student_id, name, status, time_str, subject or '-', period or '-'])
    
    output.seek(0)
    return send_file(output, mimetype='text/csv', as_attachment=True,
                    download_name=f'daily_attendance_{program}_{year}_{section}_{date}.csv')


def export_daily_excel(attendance_data, program, year, section, date):
    """Export daily attendance to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"
    
    # Headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = 'Daily Attendance Report'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = f'Program: {program}  Year: {year}  Section: {section}  Date: {date}'
    
    ws['A4'] = 'Student ID'
    ws['B4'] = 'Name'
    ws['C4'] = 'Status'
    ws['D4'] = 'Time'
    ws['E4'] = 'Subject'
    ws['F4'] = 'Period'
    
    for cell in ['A4', 'B4', 'C4', 'D4', 'E4', 'F4']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row = 5
    for record in attendance_data:
        student_id, name, status, att_date, att_time, subject, period = record
        status = status or 'No Record'
        time_str = att_time.strftime('%H:%M:%S') if hasattr(att_time, 'strftime') else str(att_time or '')
        
        ws[f'A{row}'] = student_id
        ws[f'B{row}'] = name
        ws[f'C{row}'] = status
        ws[f'D{row}'] = time_str
        ws[f'E{row}'] = subject or '-'
        ws[f'F{row}'] = period or '-'
        
        # Color code status
        if status == 'Present':
            ws[f'C{row}'].font = Font(color="00B050")
        elif status == 'Absent':
            ws[f'C{row}'].font = Font(color="FF0000")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'daily_attendance_{program}_{year}_{section}_{date}.xlsx')


def export_monthly_csv(attendance_data, program, year, month, year_val):
    """Export monthly percentage report to CSV"""
    output = BytesIO()
    writer = csv.writer(output)
    
    from calendar import month_name
    month_str = month_name[month]
    
    # Write header
    writer.writerow(['Monthly Attendance Percentage Report'])
    writer.writerow([f'Program: {program}', f'Year: {year}', f'Month: {month_str}', f'Year: {year_val}'])
    writer.writerow([])
    
    writer.writerow(['Section', 'Student ID', 'Name', 'Present Days', 'Total Days', 'Percentage (%)'])
    
    # Write data
    for record in attendance_data:
        student_id, name, section, present_days, total_days, percentage = record
        percentage_val = percentage if percentage is not None else 0
        writer.writerow([section, student_id, name, present_days or 0, total_days or 0, f'{percentage_val}%'])
    
    output.seek(0)
    return send_file(output, mimetype='text/csv', as_attachment=True,
                    download_name=f'monthly_attendance_{program}_{year}_{month_str}_{year_val}.csv')


def export_monthly_excel(attendance_data, program, year, month, year_val):
    """Export monthly percentage report to Excel"""
    from calendar import month_name
    month_str = month_name[month]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"
    
    # Headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = 'Monthly Attendance Percentage Report'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = f'Program: {program}  Year: {year}  Month: {month_str}  Year: {year_val}'
    
    ws['A4'] = 'Section'
    ws['B4'] = 'Student ID'
    ws['C4'] = 'Name'
    ws['D4'] = 'Present Days'
    ws['E4'] = 'Total Days'
    ws['F4'] = 'Percentage (%)'
    
    for cell in ['A4', 'B4', 'C4', 'D4', 'E4', 'F4']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row = 5
    for record in attendance_data:
        student_id, name, section, present_days, total_days, percentage = record
        percentage_val = percentage if percentage is not None else 0
        
        ws[f'A{row}'] = section
        ws[f'B{row}'] = student_id
        ws[f'C{row}'] = name
        ws[f'D{row}'] = present_days or 0
        ws[f'E{row}'] = total_days or 0
        ws[f'F{row}'] = percentage_val
        
        # Color code percentage
        if percentage_val >= 75:
            ws[f'F{row}'].font = Font(color="00B050")
        elif percentage_val >= 50:
            ws[f'F{row}'].font = Font(color="FFC7CE")
        else:
            ws[f'F{row}'].font = Font(color="FF0000")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'monthly_attendance_{program}_{year}_{month_str}_{year_val}.xlsx')


def export_full_csv(attendance_data):
    """Export full attendance data to CSV"""
    output = BytesIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Complete Attendance Export'])
    writer.writerow(['Generated on', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    
    writer.writerow(['Student ID', 'Name', 'Program', 'Year', 'Section', 'Date', 'Time', 'Status', 'Subject', 'Period'])
    
    # Write data
    for record in attendance_data:
        student_id, name, program, year, section, date, time, status, subject, period = record
        date_str = date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date)
        time_str = time.strftime('%H:%M:%S') if hasattr(time, 'strftime') else str(time)
        
        writer.writerow([student_id, name, program, year, section, date_str, time_str, status, subject or '-', period or '-'])
    
    output.seek(0)
    return send_file(output, mimetype='text/csv', as_attachment=True,
                    download_name=f'complete_attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')


def export_full_excel(attendance_data):
    """Export full attendance data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Data"
    
    # Headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = 'Complete Attendance Export'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = f'Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    ws['A4'] = 'Student ID'
    ws['B4'] = 'Name'
    ws['C4'] = 'Program'
    ws['D4'] = 'Year'
    ws['E4'] = 'Section'
    ws['F4'] = 'Date'
    ws['G4'] = 'Time'
    ws['H4'] = 'Status'
    ws['I4'] = 'Subject'
    ws['J4'] = 'Period'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws[f'{col}4'].fill = header_fill
        ws[f'{col}4'].font = header_font
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row = 5
    for record in attendance_data:
        student_id, name, program, year, section, date, time, status, subject, period = record
        date_str = date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date)
        time_str = time.strftime('%H:%M:%S') if hasattr(time, 'strftime') else str(time)
        
        ws[f'A{row}'] = student_id
        ws[f'B{row}'] = name
        ws[f'C{row}'] = program
        ws[f'D{row}'] = year
        ws[f'E{row}'] = section
        ws[f'F{row}'] = date_str
        ws[f'G{row}'] = time_str
        ws[f'H{row}'] = status
        ws[f'I{row}'] = subject or '-'
        ws[f'J{row}'] = period or '-'
        
        # Color code status
        if status == 'Present':
            ws[f'H{row}'].font = Font(color="00B050")
        elif status == 'Absent':
            ws[f'H{row}'].font = Font(color="FF0000")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'complete_attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')


# Logout route
@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect('/')
@app.route('/db-test')
def db_test():
    try:
        conn = connect_db()
        if conn.is_connected():
            conn.close()
            return "DATABASE CONNECTED SUCCESSFULLY"
        return "DATABASE CONNECTION FAILED", 500
    except Exception as e:
        return f"DATABASE CONNECTION ERROR: {e}", 500
    return "DATABASE CONNECTION FAILED", 500
import os


def get_ssl_context():
    """Return an SSL context for HTTPS if USE_HTTPS is enabled."""
    use_https = os.environ.get("USE_HTTPS", "false").lower() in ("1", "true", "yes", "on")
    if not use_https:
        return None

    cert_path = os.environ.get("SSL_CERT_PATH")
    key_path = os.environ.get("SSL_KEY_PATH")
    if cert_path and key_path:
        if os.path.exists(cert_path) and os.path.exists(key_path):
            return (cert_path, key_path)
        print(f"SSL cert/key files not found: {cert_path}, {key_path}")

    print("Starting Flask with adhoc SSL context. Accept the self-signed certificate in your browser.")
    return "adhoc"


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 8080))
    ssl_context = get_ssl_context()
    if ssl_context:
        app.run(host='0.0.0.0', port=port, debug=False, ssl_context=ssl_context)
    else:
       app.run(host='0.0.0.0', port=port, debug=False)